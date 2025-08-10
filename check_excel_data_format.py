#!/usr/bin/env python3
"""
Check if Excel file contains actual text with leading zeros or just formatted numbers
"""
import pandas as pd
from openpyxl import load_workbook
import sys
sys.path.insert(0, '/Users/biswambarpradhan/UpSkill/ftt-ml/backend')

def analyze_excel_file(file_path):
    """Analyze Excel file to see actual vs displayed values"""
    print(f"🔍 ANALYZING EXCEL FILE: {file_path}")
    print("=" * 60)
    
    try:
        # Method 1: Read with pandas (what our code does)
        print("📊 Method 1: pandas.read_excel() - what reconciliation sees:")
        df_pandas = pd.read_excel(file_path, dtype=str)  # Force string reading
        print(df_pandas.head())
        print(f"Data types: {df_pandas.dtypes.to_dict()}")
        
        # Method 2: Read with openpyxl directly (raw Excel values)
        print(f"\n📊 Method 2: openpyxl direct reading - actual Excel storage:")
        wb = load_workbook(file_path, data_only=True)  # Get calculated values
        ws = wb.active
        
        # Read first few rows
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx >= 5:  # Only show first 5 rows
                break
            print(f"Row {row_idx + 1}: {row}")
        
        # Method 3: Check cell formats
        print(f"\n📊 Method 3: Cell formats and types:")
        wb_formatted = load_workbook(file_path, data_only=False)  # Get formulas/formats
        ws_formatted = wb_formatted.active
        
        for row_idx in range(1, 6):  # Check first 5 rows
            for col_idx in range(1, 5):  # Check first 4 columns
                cell = ws_formatted.cell(row=row_idx, col=col_idx)
                value = cell.value
                data_type = cell.data_type
                number_format = cell.number_format
                
                if value is not None:
                    print(f"Cell {cell.coordinate}: Value={value!r}, Type={data_type}, Format='{number_format}'")
        
        return True
        
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")
        return False

def create_properly_formatted_excel():
    """Create an Excel file with properly preserved leading zeros"""
    print(f"\n📝 CREATING PROPERLY FORMATTED EXCEL FILE")
    print("=" * 50)
    
    # Create test data
    data = {
        'TransactionID': ['01', '02', '07', '10', '05'],
        'Amount': [100.0, 200.0, 700.0, 1000.0, 500.0]
    }
    df = pd.DataFrame(data)
    
    # Save with specific formatting to preserve leading zeros
    file_path = 'test_leading_zeros_preserved.xlsx'
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Get the worksheet to apply formatting
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Format TransactionID column as text
        for row in range(2, len(df) + 2):  # Skip header row
            cell = worksheet[f'A{row}']
            cell.number_format = '@'  # Text format
            
    print(f"✅ Created {file_path} with text-formatted TransactionID column")
    
    # Test reading it back
    print(f"\n🔍 Testing reading back the properly formatted file:")
    df_test = pd.read_excel(file_path, dtype={'TransactionID': str})
    print(df_test['TransactionID'].tolist())
    
    return file_path

def main():
    print("EXCEL DATA FORMAT ANALYSIS")
    print("=" * 60)
    print("Checking if Excel files contain actual text or formatted numbers")
    print()
    
    # If you have an actual Excel file, replace this path
    test_file_path = None
    
    # Ask user for file path
    print("📝 To analyze your actual Excel file:")
    print("   1. Put your Excel file in this directory")
    print("   2. Update the file path below")
    print()
    
    # You can uncomment and modify this line to test your actual file
    # test_file_path = "your_excel_file.xlsx"
    
    if test_file_path:
        analyze_excel_file(test_file_path)
    else:
        print("⚠️  No Excel file path provided for analysis")
    
    # Create a properly formatted example
    created_file = create_properly_formatted_excel()
    
    print(f"\n{'='*60}")
    print("🎯 EXCEL LEADING ZERO SUMMARY")
    print("=" * 60)
    
    print("🔍 COMMON EXCEL ISSUES WITH LEADING ZEROS:")
    print("   1. Excel auto-converts '01' to 1 when entered as number")
    print("   2. Display formatting (00) shows '01' but stores 1")
    print("   3. Text format (@) preserves actual '01' as text")
    
    print(f"\n✅ SOLUTIONS:")
    print("   1. Format cells as TEXT before entering data")
    print("   2. Use apostrophe prefix: '01 (forces text)")
    print("   3. Import CSV files instead (preserves text)")
    print("   4. Use XLSX files with proper text formatting")
    
    print(f"\n📝 TO FIX YOUR EXCEL FILES:")
    print("   1. Open Excel file")
    print("   2. Select columns with leading zeros (like TransactionID)")
    print("   3. Format > Cells > Number > Text")
    print("   4. Re-enter the data with leading zeros")
    print("   5. Save the file")
    
    print(f"\n🔧 TO TEST YOUR FILE:")
    print(f"   1. Replace 'test_file_path' with your Excel file path")
    print(f"   2. Run this script again to see the analysis")

if __name__ == "__main__":
    main()