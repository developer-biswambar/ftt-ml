#!/usr/bin/env python3
"""
Debug Excel files with apostrophe prefix ('01) 
Excel uses apostrophes to force text format, but pandas might handle them differently
"""
import sys
import pandas as pd
from io import BytesIO
import numpy as np
from openpyxl import Workbook

# Add backend to path
sys.path.insert(0, '/Users/biswambarpradhan/UpSkill/ftt-ml/backend')

def create_excel_with_apostrophe_prefix():
    """Create Excel file with apostrophe-prefixed values like '01"""
    print("📊 CREATING EXCEL FILE WITH APOSTROPHE PREFIX ('01)")
    print("=" * 60)
    
    # Create workbook with apostrophe-prefixed values
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add headers
    headers = ['TransactionID', 'Amount', 'Status']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data with apostrophe prefix (how Excel stores '01)
    data_rows = [
        ["'01", 100.0, "Active"],
        ["'02", 200.0, "Pending"], 
        ["'07", 700.0, "Active"],
        ["'10", 1000.0, "Complete"],
        ["'05", 500.0, "Active"]
    ]
    
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to BytesIO for testing
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    print("✅ Created Excel with apostrophe-prefixed values:")
    print("   '01, '02, '07, '10, '05")
    print()
    
    return excel_buffer

def test_pandas_reading_apostrophe():
    """Test how pandas reads Excel files with apostrophe prefix"""
    print("🔍 TESTING PANDAS READING OF APOSTROPHE-PREFIXED VALUES")
    print("=" * 60)
    
    excel_buffer = create_excel_with_apostrophe_prefix()
    
    # Test 1: Standard pandas reading
    excel_buffer.seek(0)
    df_standard = pd.read_excel(excel_buffer, engine='openpyxl')
    
    print("📖 Standard pandas.read_excel():")
    print(f"   Shape: {df_standard.shape}")
    for col in df_standard.columns:
        sample_vals = df_standard[col].head().tolist()
        dtype = df_standard[col].dtype
        print(f"   {col}: {sample_vals} (dtype: {dtype})")
    
    # Test 2: Force string reading
    excel_buffer.seek(0)
    df_string = pd.read_excel(excel_buffer, engine='openpyxl', dtype=str)
    
    print(f"\n📖 pandas.read_excel(dtype=str):")
    for col in df_string.columns:
        sample_vals = df_string[col].head().tolist()
        dtype = df_string[col].dtype
        print(f"   {col}: {sample_vals} (dtype: {dtype})")
    
    # Test 3: Detailed value inspection
    print(f"\n🔍 DETAILED VALUE INSPECTION:")
    print("=" * 40)
    
    for i in range(min(3, len(df_standard))):
        standard_val = df_standard['TransactionID'].iloc[i]
        string_val = df_string['TransactionID'].iloc[i]
        
        print(f"\nRow {i}:")
        print(f"   Standard reading: {standard_val!r} (type: {type(standard_val).__name__})")
        print(f"   String reading: {string_val!r} (type: {type(string_val).__name__})")
        
        # Check for apostrophe
        if isinstance(string_val, str):
            has_apostrophe = string_val.startswith("'")
            without_apostrophe = string_val.lstrip("'")
            print(f"   Has apostrophe: {has_apostrophe}")
            print(f"   Without apostrophe: {without_apostrophe!r}")
            print(f"   Length: {len(string_val)}")
            
            # Check ASCII codes
            print(f"   ASCII codes: {[ord(c) for c in string_val]}")
    
    return df_standard, df_string

def test_reconciliation_with_apostrophe():
    """Test reconciliation service with apostrophe-prefixed Excel data"""
    print(f"\n🔄 TESTING RECONCILIATION WITH APOSTROPHE-PREFIXED DATA")
    print("=" * 60)
    
    try:
        from app.services.reconciliation_service import OptimizedFileProcessor
        from app.models.recon_models import ReconciliationRule
        
        # Create two Excel files with apostrophe-prefixed data
        excel_a = create_excel_with_apostrophe_prefix()
        excel_b = create_excel_with_apostrophe_prefix()  # Same data for matching
        
        # Create mock upload files
        class MockUploadFile:
            def __init__(self, content: BytesIO, filename: str):
                self.file = content
                self.filename = filename
        
        mock_file_a = MockUploadFile(excel_a, 'file_a.xlsx')
        mock_file_b = MockUploadFile(excel_b, 'file_b.xlsx')
        
        processor = OptimizedFileProcessor()
        
        # Read files through reconciliation service
        print("📖 Reading files through reconciliation service...")
        df_a = processor.read_file(mock_file_a)
        df_b = processor.read_file(mock_file_b)
        
        print(f"\n📊 File A after reconciliation service:")
        for col in df_a.columns:
            sample_vals = df_a[col].head(3).tolist()
            dtype = df_a[col].dtype
            print(f"   {col}: {sample_vals} (dtype: {dtype})")
        
        print(f"\n📊 File B after reconciliation service:")
        for col in df_b.columns:
            sample_vals = df_b[col].head(3).tolist() 
            dtype = df_b[col].dtype
            print(f"   {col}: {sample_vals} (dtype: {dtype})")
        
        # Test matching manually
        print(f"\n🔍 MANUAL MATCHING TEST:")
        print("=" * 30)
        
        for i in range(min(3, len(df_a))):
            val_a = df_a['TransactionID'].iloc[i]
            val_b = df_b['TransactionID'].iloc[i]
            
            print(f"\nRow {i}:")
            print(f"   File A: {val_a!r}")
            print(f"   File B: {val_b!r}")
            print(f"   Direct equality: {val_a == val_b}")
            print(f"   String equality: {str(val_a) == str(val_b)}")
            
            # Test reconciliation equals method
            try:
                equals_result = processor._check_equals_match(val_a, val_b)
                print(f"   _check_equals_match: {equals_result}")
            except Exception as e:
                print(f"   _check_equals_match ERROR: {e}")
                
            # Test with apostrophe stripped
            if isinstance(val_a, str) and isinstance(val_b, str):
                stripped_a = val_a.lstrip("'")
                stripped_b = val_b.lstrip("'")
                print(f"   Stripped A: {stripped_a!r}")
                print(f"   Stripped B: {stripped_b!r}")
                print(f"   Stripped equality: {stripped_a == stripped_b}")
        
        # Run actual reconciliation
        print(f"\n🔍 RUNNING ACTUAL RECONCILIATION:")
        recon_rules = [
            ReconciliationRule(
                LeftFileColumn="TransactionID",
                RightFileColumn="TransactionID",
                MatchType="equals"
            )
        ]
        
        results = processor.reconcile_files_optimized(df_a, df_b, recon_rules)
        
        matched_count = len(results['matched'])
        unmatched_a_count = len(results['unmatched_file_a'])
        unmatched_b_count = len(results['unmatched_file_b'])
        
        print(f"   📈 Total A: {len(df_a)}, Total B: {len(df_b)}")
        print(f"   📈 Matched: {matched_count}")
        print(f"   📈 Unmatched A: {unmatched_a_count}")
        print(f"   📈 Unmatched B: {unmatched_b_count}")
        print(f"   📈 Expected: 5 matches")
        
        if matched_count < 5:
            print(f"   ❌ ISSUE CONFIRMED: Apostrophe prefix causing match failures")
        else:
            print(f"   ✅ SUCCESS: All records matched despite apostrophe prefix")
        
        return results, matched_count
        
    except Exception as e:
        print(f"❌ Error testing reconciliation: {e}")
        import traceback
        traceback.print_exc()
        return None, 0

def suggest_fixes():
    """Suggest fixes for apostrophe prefix issues"""
    print(f"\n🔧 SUGGESTED FIXES FOR APOSTROPHE PREFIX ISSUE")
    print("=" * 60)
    
    print("🎯 ROOT CAUSE:")
    print("   Excel uses apostrophe prefix ('01) to force text format")
    print("   Pandas reads this as literal string with apostrophe")
    print("   Reconciliation compares '01 vs '02 (includes apostrophe)")
    
    print(f"\n✅ SOLUTION 1: Strip apostrophe in reconciliation matching")
    print("   Modify _check_equals_match to handle apostrophe prefix")
    
    print(f"\n✅ SOLUTION 2: Clean data during file reading") 
    print("   Strip leading apostrophes from Excel text columns")
    
    print(f"\n✅ SOLUTION 3: Fix Excel files")
    print("   Format cells as Text and re-enter without apostrophe")
    print("   Or use proper text formatting in Excel")

def main():
    print("EXCEL APOSTROPHE PREFIX DEBUG")
    print("=" * 60)
    print("Debugging Excel files with apostrophe prefix ('01)")
    print("This is how Excel forces text format for leading zeros")
    print()
    
    # Test pandas reading behavior
    df_standard, df_string = test_pandas_reading_apostrophe()
    
    # Test reconciliation with apostrophe data
    results, matched_count = test_reconciliation_with_apostrophe()
    
    # Suggest fixes
    suggest_fixes()
    
    print(f"\n{'='*60}")
    print("🎯 APOSTROPHE PREFIX SUMMARY")
    print("=" * 60)
    
    if matched_count < 5:
        print("❌ ISSUE CONFIRMED:")
        print("   Apostrophe prefix ('01) is preventing matches")
        print("   Manual verification shows '01 = '01 but reconciliation fails")
        
        print(f"\n🔧 IMMEDIATE FIX NEEDED:")
        print("   Update reconciliation matching to handle apostrophe prefix")
        print("   Or clean apostrophes during Excel file reading")
        
    else:
        print("✅ NO ISSUE FOUND:")
        print("   Reconciliation handled apostrophe prefix correctly")
        print("   The issue might be elsewhere")
    
    print(f"\n{'='*60}")

if __name__ == "__main__":
    main()